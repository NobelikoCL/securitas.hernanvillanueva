from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy
from django.contrib import messages
from django.db import transaction
import pandas as pd
from io import BytesIO
import os
from django.core.files.base import ContentFile
from django.http import HttpResponse, FileResponse, Http404
from django.conf import settings
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Vehiculo, Personal, Instalacion, Turno, AsignacionVehiculo, AsignacionInstalacion, GestionTransporte, CentroCosto, Cliente
from .forms import VehiculoForm, PersonalForm, FileUploadForm, ClienteForm, CentroCostoForm, InstalacionForm

class VehiculoListView(ListView):
    model = Vehiculo
    template_name = 'flota_app/vehiculo_list.html'
    context_object_name = 'vehiculos'

    def get_queryset(self):
        # Optimizar para obtener información relacionada
        return Vehiculo.objects.all().select_related('instalacion_base__centro_costo__cliente')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Vehículos"
        # Se podría añadir un selector de cliente para filtrar la lista de vehículos si fuera necesario
        # context['clientes'] = Cliente.objects.all()
        return context

class VehiculoCreateView(CreateView):
    model = Vehiculo
    form_class = VehiculoForm
    template_name = 'flota_app/vehiculo_form.html'
    success_url = reverse_lazy('flota_app:vehiculo_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Añadir Vehículo"
        return context

class VehiculoUpdateView(UpdateView):
    model = Vehiculo
    form_class = VehiculoForm
    template_name = 'flota_app/vehiculo_form.html'
    success_url = reverse_lazy('flota_app:vehiculo_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Editar Vehículo"
        return context

class VehiculoDeleteView(DeleteView):
    model = Vehiculo
    template_name = 'flota_app/vehiculo_confirm_delete.html'
    success_url = reverse_lazy('flota_app:vehiculo_list')
    context_object_name = 'vehiculo' # Para usar {{ vehiculo }} en la plantilla

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Confirmar Eliminación de Vehículo"
        return context

# Más vistas se añadirán aquí para otras funcionalidades.

def dashboard_view(request):
    context = {
        'titulo_pagina': "Dashboard Principal"
    }
    return render(request, 'flota_app/dashboard.html', context)

# Vistas para Personal
class PersonalListView(ListView):
    model = Personal
    template_name = 'flota_app/personal_list.html'
    context_object_name = 'personal_list' # Cambiado para evitar conflicto con 'personal' si se usa en el contexto

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Listado de Personal"
        return context

class PersonalDetailView(DetailView):
    model = Personal
    template_name = 'flota_app/personal_detail.html'
    context_object_name = 'personal'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Detalle de {self.object.nombre} {self.object.apellido}"
        return context

class PersonalCreateView(CreateView):
    model = Personal
    form_class = PersonalForm
    template_name = 'flota_app/personal_form.html'
    success_url = reverse_lazy('flota_app:personal_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Añadir Personal"
        context['es_creacion'] = True # Para diferenciar en la plantilla si es necesario
        return context

class PersonalUpdateView(UpdateView):
    model = Personal
    form_class = PersonalForm
    template_name = 'flota_app/personal_form.html'
    success_url = reverse_lazy('flota_app:personal_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar Personal: {self.object.nombre} {self.object.apellido}"
        context['es_creacion'] = False
        return context

class PersonalDeleteView(DeleteView):
    model = Personal
    template_name = 'flota_app/personal_confirm_delete.html'
    success_url = reverse_lazy('flota_app:personal_list')
    context_object_name = 'personal'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Confirmar Eliminación de {self.object.nombre} {self.object.apellido}"
        return context

import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from django.views import View
from .forms import FileUploadForm
from .models import Instalacion, CentroCosto # Asegúrate que User esté importado si lo usas directamente
from django.contrib.auth.models import User
from datetime import datetime

class PersonalBulkUploadView(View):
    template_name = 'flota_app/personal_bulk_upload.html'

    def get(self, request, *args, **kwargs):
        form = FileUploadForm()
        
        # Crear plantilla de ejemplo si no existe
        template_path = os.path.join(settings.MEDIA_ROOT, 'plantillas', 'plantilla_carga_masiva_personal.xlsx')
        os.makedirs(os.path.dirname(template_path), exist_ok=True)
        
        # Si el archivo no existe o se solicita regenerar
        if not os.path.exists(template_path) or request.GET.get('regenerar'):
            self.generar_plantilla_ejemplo(template_path)
        
        context = {
            'form': form,
            'titulo_pagina': "Carga Masiva de Personal",
            'ejemplo_archivo': os.path.join(settings.MEDIA_URL, 'plantillas/plantilla_carga_masiva_personal.xlsx')
        }
        return render(request, self.template_name, context)
    
    def generar_plantilla_ejemplo(self, filepath):
        """Genera un archivo Excel de ejemplo para la carga masiva de personal"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Personal"
        
        # Estilos para la cabecera
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Columnas requeridas
        headers = [
            "nombre", "apellido", "rut", "fecha_nacimiento", "direccion",
            "telefono_contacto", "acepta_mensajeria_whatsapp", "legajo",
            "fecha_ingreso", "instalacion_trabajo", "centro_costo"
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
            # Ajustar ancho de columna
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(header) + 2, 15)
        
        # Agregar fila de ejemplo
        ejemplo = [
            "Juan", "Pérez González", "12.345.678-9", "1985-05-15", "Av. Principal 1234",
            "912345678", "SI", "EMP-001", "2023-01-15", "Oficina Central", "CC-001"
        ]
        
        for col_num, value in enumerate(ejemplo, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            if headers[col_num-1] in ['fecha_nacimiento', 'fecha_ingreso']:
                cell.number_format = 'YYYY-MM-DD'
        
        # Agregar validación de datos
        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"SI,NO"',
            showErrorMessage=True,
            errorTitle='Valor no válido',
            error='Debe ser SI o NO',
            showDropDown=True
        )
        dv.add(f'G2:G1048576')  # Para la columna acepta_mensajeria_whatsapp
        ws.add_data_validation(dv)
        
        # Guardar el archivo
        wb.save(filepath)
        return filepath

    def post(self, request, *args, **kwargs):
        form = FileUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
            return self.get(request, *args, **kwargs)
            
        file = request.FILES['file']
        file_extension = os.path.splitext(file.name)[1].lower()
        
        try:
            # Leer el archivo según su extensión
            if file_extension == '.csv':
                df = pd.read_csv(file, dtype=str, encoding='utf-8')
            elif file_extension in ['.xls', '.xlsx']:
                df = pd.read_excel(file, dtype=str, engine='openpyxl')
            else:
                messages.error(request, "Formato de archivo no soportado. Use archivos CSV o Excel (XLSX).")
                return redirect('flota_app:personal_bulk_upload')

            # Eliminar filas completamente vacías
            df = df.dropna(how='all')
            
            # Verificar si el archivo está vacío
            if df.empty:
                messages.error(request, "El archivo está vacío o no contiene datos válidos.")
                return redirect('flota_app:personal_bulk_upload')

            # Normalizar nombres de columnas (quitar espacios, a minúsculas, sin caracteres especiales)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_').str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8')
            
            # Verificar columnas requeridas
            required_columns = [
                'nombre', 'apellido', 'rut', 'fecha_nacimiento', 
                'direccion', 'telefono_contacto', 'acepta_mensajeria_whatsapp', 
                'legajo', 'fecha_ingreso'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(
                    request, 
                    f"Faltan las siguientes columnas obligatorias: {', '.join(missing_columns)}. "
                    f"Por favor, descargue la plantilla de ejemplo para ver el formato correcto."
                )
                return redirect('flota_app:personal_bulk_upload')

            errors = []
            success_count = 0
            
            # Listas para almacenar datos de éxito y advertencias
            success_list = []
            warning_list = []
            
            # Obtener todas las instalaciones y centros de costo para validación
            instalaciones = {inst.nombre.lower().strip(): inst for inst in Instalacion.objects.all()}
            centros_costo = {str(cc.codigo).lower().strip(): cc for cc in CentroCosto.objects.all()}
            
            # Validar cada fila del archivo
            for index, row in df.iterrows():
                row_errors = []
                row_num = index + 2  # +2 porque Excel comienza en 1 y la primera fila es el encabezado
                
                # Validar campos obligatorios
                for col in required_columns:
                    if pd.isna(row.get(col, '')) or str(row[col]).strip() == '':
                        row_errors.append(f"La columna '{col}' es obligatoria")
                
                # Validar formato de RUT
                rut = str(row.get('rut', '')).strip()
                if rut:
                    # Validar formato básico de RUT (acepta con o sin puntos/guion)
                    import re
                    if not re.match(r'^\d{1,3}(?:\.?\d{3}){2}-[\dkK]$', rut):
                        row_errors.append("Formato de RUT inválido. Debe ser como 12.345.678-9 o 12345678-9")
                
                # Validar fechas
                for date_col in ['fecha_nacimiento', 'fecha_ingreso']:
                    date_str = str(row.get(date_col, '')).strip()
                    if date_str:
                        try:
                            datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            row_errors.append(f"Formato de fecha inválido en {date_col}. Use YYYY-MM-DD")
                
                # Validar teléfono
                telefono = str(row.get('telefono_contacto', '')).strip()
                if telefono and not telefono.isdigit():
                    row_errors.append("El teléfono debe contener solo números")
                
                # Validar SI/NO para WhatsApp
                whatsapp = str(row.get('acepta_mensajeria_whatsapp', '')).strip().upper()
                if whatsapp not in ['SI', 'NO']:
                    row_errors.append("El campo 'acepta_mensajeria_whatsapp' debe ser 'SI' o 'NO'")
                
                # Validar instalación de trabajo
                instalacion_nombre = str(row.get('instalacion_trabajo', '')).strip().lower()
                if instalacion_nombre and instalacion_nombre not in instalaciones:
                    row_errors.append(f"Instalación de trabajo no encontrada: {row.get('instalacion_trabajo')}")
                
                # Validar centro de costo
                centro_costo_codigo = str(row.get('centro_costo', '')).strip().lower()
                if centro_costo_codigo and centro_costo_codigo not in centros_costo:
                    row_errors.append(f"Centro de costo no encontrado: {row.get('centro_costo')}")
                
                if row_errors:
                    errors.append({
                        'fila': row_num,
                        'errores': row_errors,
                        'datos': row.to_dict()
                    })
                    continue
                
                # Si llegamos aquí, la fila es válida
                try:
                    with transaction.atomic():
                        # Crear o actualizar el personal
                        personal_data = {
                            'nombre': str(row['nombre']).strip(),
                            'apellido': str(row['apellido']).strip(),
                            'rut': rut.replace('.', '').replace('-', '').upper(),
                            'fecha_nacimiento': datetime.strptime(str(row['fecha_nacimiento']).strip(), '%Y-%m-%d').date(),
                            'direccion': str(row['direccion']).strip(),
                            'telefono_contacto': telefono,
                            'acepta_mensajeria_whatsapp': whatsapp == 'SI',
                            'legajo': str(row['legajo']).strip(),
                            'fecha_ingreso': datetime.strptime(str(row['fecha_ingreso']).strip(), '%Y-%m-%d').date(),
                        }
                        
                        # Asignar instalación si existe
                        if instalacion_nombre:
                            personal_data['instalacion_trabajo'] = instalaciones[instalacion_nombre]
                        
                        # Asignar centro de costo si existe
                        if centro_costo_codigo:
                            personal_data['centro_costo'] = centros_costo[centro_costo_codigo]
                        
                        # Verificar si ya existe un personal con el mismo RUT o legajo
                        personal_existente = Personal.objects.filter(
                            rut=personal_data['rut']
                        ).first()
                        
                        if personal_existente:
                            # Actualizar registro existente
                            for key, value in personal_data.items():
                                setattr(personal_existente, key, value)
                            personal_existente.save()
                            success_list.append(f"Fila {row_num}: Actualizado {personal_data['nombre']} {personal_data['apellido']} (RUT: {personal_data['rut']})")
                        else:
                            # Crear nuevo registro
                            personal = Personal.objects.create(**personal_data)
                            success_list.append(f"Fila {row_num}: Creado {personal.nombre} {personal.apellido} (RUT: {personal.rut})")
                        
                        success_count += 1
                        
                except Exception as e:
                    error_msg = f"Error en fila {row_num}: {str(e)}"
                    errors.append({
                        'fila': row_num,
                        'errores': [error_msg],
                        'datos': row.to_dict()
                    })
            
            # Preparar mensajes de resultado
            if success_count > 0:
                success_message = f"Se procesaron correctamente {success_count} registros."
                if len(success_list) > 5:
                    success_message += "<br>" + "<br>".join(success_list[:5])
                    success_message += f"<br>... y {len(success_list) - 5} más"
                else:
                    success_message += "<br>" + "<br>".join(success_list)
                messages.success(request, success_message, extra_tags='safe')
            
            if errors:
                error_message = f"Se encontraron {len(errors)} errores en el archivo:"
                for i, error in enumerate(errors[:5], 1):
                    error_message += f"<br>{i}. Fila {error['fila']}: {', '.join(error['errores'])}"
                if len(errors) > 5:
                    error_message += f"<br>... y {len(errors) - 5} errores más"
                messages.error(request, error_message, extra_tags='safe')
            
            return redirect('flota_app:personal_bulk_upload')
            
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            return redirect('flota_app:personal_bulk_upload')

        # Si el formulario no es válido, mostrar errores
        messages.error(request, "Por favor, corrija los errores en el formulario.")
        return redirect('flota_app:personal_bulk_upload')

# Vistas para Cliente, Centro de Costo e Instalaciones

# Vistas CRUD para Cliente
class ClienteListView(ListView):
    model = Cliente
    template_name = 'flota_app/cliente_list.html' # Nueva plantilla para listar clientes
    context_object_name = 'clientes'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Gestión de Clientes"
        return context

class ClienteCreateView(CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'flota_app/cliente_form.html' # Reutilizar o adaptar la existente
    success_url = reverse_lazy('flota_app:cliente_list')

    def form_valid(self, form):
        messages.success(self.request, f"Cliente '{form.instance.nombre}' creado correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Registrar Nuevo Cliente"
        context['es_creacion'] = True
        return context

class ClienteUpdateView(UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'flota_app/cliente_form.html' # Reutilizar o adaptar la existente
    success_url = reverse_lazy('flota_app:cliente_list')

    def form_valid(self, form):
        messages.success(self.request, f"Cliente '{form.instance.nombre}' actualizado correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar Cliente: {self.object.nombre}"
        context['es_creacion'] = False
        return context

class ClienteDeleteView(DeleteView):
    model = Cliente
    template_name = 'flota_app/cliente_confirm_delete.html' # Nueva plantilla
    success_url = reverse_lazy('flota_app:cliente_list')
    context_object_name = 'cliente'

    def form_valid(self, form):
        # Verificar si hay Centros de Costo asociados antes de eliminar
        if self.object.centros_costo.exists():
            messages.error(self.request, f"No se puede eliminar el cliente '{self.object.nombre}' porque tiene Centros de Costo asociados. Por favor, elimínelos o reasígnelos primero.")
            # Redirigir a la lista de clientes o a la vista de detalle del cliente
            return redirect('flota_app:cliente_list') # O self.object.get_absolute_url() si está definido

        messages.success(self.request, f"Cliente '{self.object.nombre}' eliminado correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Confirmar Eliminación de Cliente: {self.object.nombre}"
        return context

# Vistas para CentroCosto
class CentroCostoListView(ListView):
    model = CentroCosto
    template_name = 'flota_app/centros_costo/centrocosto_list.html'
    context_object_name = 'centros_costo'

    def dispatch(self, request, *args, **kwargs):
        self.cliente = get_object_or_404(Cliente, pk=self.kwargs.get('cliente_id'))
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Filtrar por el cliente obtenido en dispatch
        return CentroCosto.objects.filter(cliente=self.cliente)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Centros de Costo de: {self.cliente.nombre}"
        context['cliente'] = self.cliente
        return context

class CentroCostoCreateView(CreateView):
    model = CentroCosto
    form_class = CentroCostoForm
    template_name = 'flota_app/centros_costo/centrocosto_form.html'
    # success_url se definirá en get_success_url

    def dispatch(self, request, *args, **kwargs):
        self.cliente = get_object_or_404(Cliente, pk=self.kwargs.get('cliente_id'))
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['cliente_obj'] = self.cliente # Pasar el cliente al formulario
        return kwargs

    def form_valid(self, form):
        form.instance.cliente = self.cliente # Asignar el cliente a la instancia del CC
        messages.success(self.request, f"Centro de Costo '{form.instance.nombre}' creado para {self.cliente.nombre}.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Añadir Centro de Costo para: {self.cliente.nombre}"
        context['cliente'] = self.cliente
        context['es_creacion'] = True
        return context

    def get_success_url(self):
        return reverse_lazy('flota_app:centrocosto_list_by_cliente', kwargs={'cliente_id': self.cliente.pk})

class CentroCostoUpdateView(UpdateView):
    model = CentroCosto
    form_class = CentroCostoForm
    template_name = 'flota_app/centros_costo/centrocosto_form.html'
    # success_url se definirá en get_success_url

    def get_object(self, queryset=None):
        # Asegurarse que el CC que se edita pertenece al cliente en contexto (si se proporciona)
        # o simplemente obtener el objeto por su pk si no hay contexto de cliente en la URL.
        # Por ahora, la URL no tiene cliente_id para update, así que solo usamos pk.
        # La lógica del form ya deshabilita el cambio de cliente.
        return super().get_object(queryset)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Pasar el cliente de la instancia actual al formulario para que lo deshabilite
        if self.object and self.object.cliente:
            kwargs['cliente_obj'] = self.object.cliente
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Centro de Costo '{form.instance.nombre}' actualizado.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar Centro de Costo: {self.object.nombre}"
        # El cliente se obtiene del objeto mismo (self.object.cliente)
        context['cliente'] = self.object.cliente
        context['es_creacion'] = False
        return context

    def get_success_url(self):
        # Redirigir a la lista de CCs del cliente al que pertenece el CC editado
        return reverse_lazy('flota_app:centrocosto_list_by_cliente', kwargs={'cliente_id': self.object.cliente.pk})

class CentroCostoDeleteView(DeleteView):
    model = CentroCosto
    template_name = 'flota_app/centros_costo/centrocosto_confirm_delete.html'
    context_object_name = 'centrocosto'
    # success_url se definirá en get_success_url

    def get_object(self, queryset=None):
        # Similar a UpdateView, por ahora solo usa pk.
        # Se podría añadir verificación de pertenencia a un cliente si la URL de borrado se anida.
        return super().get_object(queryset)

    def form_valid(self, form):
        cliente_pk = self.object.cliente.pk # Guardar pk antes de borrar
        # Verificar si hay Instalaciones asociadas
        if self.object.instalaciones.exists():
            messages.error(self.request, f"No se puede eliminar el Centro de Costo '{self.object.nombre}' porque tiene Instalaciones asociadas. Por favor, elimínelas o reasígnelas primero.")
            return redirect('flota_app:centrocosto_list_by_cliente', cliente_id=cliente_pk)

        messages.success(self.request, f"Centro de Costo '{self.object.nombre}' eliminado correctamente.")
        # No llamar a super().form_valid(form) aquí, sino realizar la eliminación después de la verificación
        self.object.delete()
        return redirect(self.get_success_url(cliente_pk=cliente_pk))

    def post(self, request, *args, **kwargs):
        # Sobrescribir post para manejar la redirección con success_url que depende del objeto
        self.object = self.get_object()
        # La lógica de form_valid ahora maneja la eliminación y redirección
        # Simplemente llamamos a form_valid (aunque no hay form real aquí)
        # Esto es un poco atípico para DeleteView, se podría hacer directamente
        if self.object.instalaciones.exists():
            messages.error(self.request, f"No se puede eliminar el Centro de Costo '{self.object.nombre}' porque tiene Instalaciones asociadas.")
            return redirect('flota_app:centrocosto_list_by_cliente', cliente_id=self.object.cliente.pk)

        nombre_cc_eliminado = self.object.nombre
        cliente_pk_redirect = self.object.cliente.pk
        self.object.delete()
        messages.success(self.request, f"Centro de Costo '{nombre_cc_eliminado}' eliminado correctamente.")
        return redirect(self.get_success_url(cliente_pk=cliente_pk_redirect))

    def get_success_url(self, cliente_pk=None):
        # Redirigir a la lista de CCs del cliente al que pertenecía el CC eliminado
        if cliente_pk:
             return reverse_lazy('flota_app:centrocosto_list_by_cliente', kwargs={'cliente_id': cliente_pk})
        # Fallback, aunque siempre debería haber un cliente_pk si el objeto existe
        return reverse_lazy('flota_app:cliente_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Confirmar Eliminación de CC: {self.object.nombre}"
        context['cliente'] = self.object.cliente # Para la plantilla
        return context

# Vistas para Instalacion
class InstalacionListView(ListView):
    model = Instalacion
    template_name = 'flota_app/instalaciones/instalacion_list.html'
    context_object_name = 'instalaciones'

    def dispatch(self, request, *args, **kwargs):
        self.cliente = get_object_or_404(Cliente, pk=self.kwargs.get('cliente_id'))
        # Centro de costo es opcional en la URL para la lista, puede venir como kwarg o GET param
        self.centro_costo = None
        centro_costo_id_kwarg = self.kwargs.get('centro_costo_id')
        centro_costo_id_get = self.request.GET.get('centro_costo_id_filter') # Desde el filtro del formulario

        if centro_costo_id_kwarg:
            self.centro_costo = get_object_or_404(CentroCosto, pk=centro_costo_id_kwarg, cliente=self.cliente)
        elif centro_costo_id_get:
             try:
                self.centro_costo = CentroCosto.objects.get(pk=centro_costo_id_get, cliente=self.cliente)
             except CentroCosto.DoesNotExist:
                # No hacer un 404 si el filtro GET es inválido, simplemente no filtrar por CC.
                pass
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = Instalacion.objects.filter(centro_costo__cliente=self.cliente)
        if self.centro_costo:
            queryset = queryset.filter(centro_costo=self.centro_costo)
        return queryset.select_related('centro_costo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        titulo = f"Instalaciones de: {self.cliente.nombre}"
        if self.centro_costo:
            titulo += f" (CC: {self.centro_costo.nombre})"
        context['titulo_pagina'] = titulo
        context['cliente'] = self.cliente
        context['centro_costo_filtrado'] = self.centro_costo # Para el filtro y otros elementos de la plantilla
        context['centros_costo_disponibles'] = CentroCosto.objects.filter(cliente=self.cliente)
        return context

class InstalacionCreateView(CreateView):
    model = Instalacion
    form_class = InstalacionForm
    template_name = 'flota_app/instalaciones/instalacion_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.cliente = get_object_or_404(Cliente, pk=self.kwargs.get('cliente_id'))
        self.centro_costo = get_object_or_404(CentroCosto, pk=self.kwargs.get('centro_costo_id'), cliente=self.cliente)
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['cliente_obj'] = self.cliente # Para que InstalacionForm filtre los CCs a este cliente
        # No es necesario pasar centro_costo_obj al form, ya que se asigna en form_valid.
        # Pero sí es útil para preseleccionar el CC en el form si es relevante.
        return kwargs

    def form_valid(self, form):
        form.instance.centro_costo = self.centro_costo # Asignar el CC del contexto de la URL
        messages.success(self.request, f"Instalación '{form.instance.nombre}' creada para CC '{self.centro_costo.nombre}'.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Añadir Instalación a CC '{self.centro_costo.nombre}' (Cliente: {self.cliente.nombre})"
        context['cliente'] = self.cliente
        context['centro_costo_actual'] = self.centro_costo # Para la plantilla
        context['es_creacion'] = True

        # Preseleccionar el centro_costo en el formulario
        form = context.get('form', self.form_class(**self.get_form_kwargs())) # Obtener form con kwargs correctos
        form.fields['centro_costo'].initial = self.centro_costo
        # Opcionalmente deshabilitar si solo se puede crear para este CC
        # form.fields['centro_costo'].widget.attrs['disabled'] = True
        context['form'] = form
        return context

    def get_success_url(self):
        return reverse_lazy('flota_app:instalacion_list_by_cc', kwargs={
            'cliente_id': self.cliente.pk,
            'centro_costo_id': self.centro_costo.pk
        })

class InstalacionUpdateView(UpdateView):
    model = Instalacion
    form_class = InstalacionForm
    template_name = 'flota_app/instalaciones/instalacion_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Pasar el cliente del CC de la instancia actual al formulario
        if self.object and self.object.centro_costo:
            kwargs['cliente_obj'] = self.object.centro_costo.cliente
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f"Instalación '{form.instance.nombre}' actualizada.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar Instalación: {self.object.nombre}"
        context['cliente'] = self.object.centro_costo.cliente # Para la plantilla
        context['centro_costo_actual'] = self.object.centro_costo # Para la plantilla
        context['es_creacion'] = False
        return context

    def get_success_url(self):
        # Redirigir a la lista de instalaciones del CC al que pertenece la instalación
        return reverse_lazy('flota_app:instalacion_list_by_cc', kwargs={
            'cliente_id': self.object.centro_costo.cliente.pk,
            'centro_costo_id': self.object.centro_costo.pk
        })

class InstalacionDeleteView(DeleteView):
    model = Instalacion
    template_name = 'flota_app/instalaciones/instalacion_confirm_delete.html'
    context_object_name = 'instalacion'

    def get_success_url(self):
        # Redirigir a la lista de instalaciones del CC al que pertenecía la instalación
        if self.object and self.object.centro_costo:
             return reverse_lazy('flota_app:instalacion_list_by_cc', kwargs={
                'cliente_id': self.object.centro_costo.cliente.pk,
                'centro_costo_id': self.object.centro_costo.pk
            })
        # Fallback (aunque no debería ocurrir si el objeto existe)
        return reverse_lazy('flota_app:cliente_list')


    def form_valid(self, form):
        messages.success(self.request, f"Instalación '{self.object.nombre}' eliminada correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Confirmar Eliminación de Instalación: {self.object.nombre}"
        context['cliente'] = self.object.centro_costo.cliente # Para la plantilla
        context['centro_costo_actual'] = self.object.centro_costo # Para la plantilla
        return context
